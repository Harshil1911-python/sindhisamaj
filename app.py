"""
Sindhi Samaj Pariwar Surat - Marriage Bureau Software
Flask backend. All templates/static files are loose in this same folder
(no subfolders), as requested. Uses SQLite as the database.
"""

import os
import io
import csv
import re
import uuid
import sqlite3
import secrets
from datetime import datetime, date
from functools import wraps

from flask import (
    Flask, request, jsonify, session, send_file, send_from_directory,
    redirect, url_for, g, abort, Response
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# DATA_DIR can be pointed at a Render persistent disk mount so the database
# and uploaded photos survive redeploys. Defaults to the app folder itself.
DATA_DIR = os.environ.get("DATA_DIR", BASE_DIR)
os.makedirs(DATA_DIR, exist_ok=True)
DB_PATH = os.path.join(DATA_DIR, "database.db")
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_IMAGE_EXT = {"png", "jpg", "jpeg", "webp", "gif"}
ALLOWED_DOC_EXT = {"pdf"}
ALLOWED_DB_EXT = {"db", "sqlite", "sqlite3"}

LOGIN_MAX_ATTEMPTS = 5
LOGIN_LOCKOUT_MINUTES = 15

app = Flask(__name__, template_folder=BASE_DIR, static_folder=BASE_DIR, static_url_path="")
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))
app.config["MAX_CONTENT_LENGTH"] = 60 * 1024 * 1024  # 60 MB uploads ceiling

# ---------------------------------------------------------------------------
# All profile fields (order matters for CSV export/import & forms)
# ---------------------------------------------------------------------------
PROFILE_FIELDS = [
    "full_name", "gender", "profile_for", "father_name", "mother_name", "guardian_name",
    "dob", "time_of_birth", "place_of_birth", "age", "height", "weight", "blood_group",
    "complexion", "body_type", "nationality", "religion", "community", "sindhi_caste",
    "sub_caste", "mother_tongue",
    "marital_status", "children",
    "occupation", "career", "company", "business", "designation", "income_monthly", "income_yearly",
    "qualification", "college", "university",
    "current_city", "native_place", "phone", "alternate_phone", "email", "whatsapp",
    "emergency_contact", "permanent_address", "temporary_address",
    "food_preference", "smoking", "drinking", "habits", "hobbies", "interests",
    "languages_known", "about_yourself",
    "expected_age", "expected_height", "expected_education", "expected_profession",
    "expected_income", "expected_location", "expected_lifestyle",
    "manglik", "rashi", "nakshatra", "mulank", "kundli_available", "birth_chart", "horoscope_notes",
    "disability", "special_notes",
    "aadhar_number", "passport_number",
]

FILE_FIELDS = ["kundli_pdf", "aadhar_photo", "passport_photo"]  # single-file inputs
GALLERY_FIELD = "photos"  # multi-file input (profile photos)


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    first_time = not os.path.exists(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    with open(os.path.join(BASE_DIR, "schema.sql"), "r") as f:
        conn.executescript(f.read())
    # migration safety net for databases created before admin_notes existed
    try:
        conn.execute("ALTER TABLE profiles ADD COLUMN admin_notes TEXT")
    except sqlite3.OperationalError:
        pass
    # create a default admin if none exists yet
    cur = conn.execute("SELECT COUNT(*) FROM admins")
    if cur.fetchone()[0] == 0:
        conn.execute(
            "INSERT INTO admins (username, password_hash) VALUES (?, ?)",
            ("admin", generate_password_hash("admin123")),
        )
        conn.commit()
        print("=" * 70)
        print(" Default admin created -> username: admin | password: admin123")
        print(" Please sign in and change this immediately.")
        print("=" * 70)
    conn.commit()
    conn.close()
    return first_time


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("admin_id"):
            if request.path.startswith("/api/") or request.path.startswith("/export") \
                    or request.path.startswith("/import") or request.path.startswith("/backup") \
                    or request.path.startswith("/restore") or request.path.startswith("/biodata"):
                return jsonify({"success": False, "message": "Not authenticated"}), 401
            return redirect("/signin.html")
        return view(*args, **kwargs)
    return wrapped


def user_login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_profile_id"):
            if request.path.startswith("/api/"):
                return jsonify({"success": False, "message": "Please log in to your profile."}), 401
            return redirect("/login.html")
        return view(*args, **kwargs)
    return wrapped


def allowed_ext(filename, allowed):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in allowed


def save_upload(file_storage, subfolder_prefix):
    if not file_storage or file_storage.filename == "":
        return None
    ext = file_storage.filename.rsplit(".", 1)[-1].lower() if "." in file_storage.filename else ""
    fname = f"{subfolder_prefix}_{uuid.uuid4().hex[:10]}.{ext}" if ext else f"{subfolder_prefix}_{uuid.uuid4().hex[:10]}"
    fname = secure_filename(fname)
    path = os.path.join(UPLOAD_DIR, fname)
    file_storage.save(path)
    return fname  # stored relative to /uploads/


def calc_age(dob_str):
    try:
        y, m, d = [int(x) for x in dob_str.split("-")]
        born = date(y, m, d)
        today = date.today()
        return today.year - born.year - ((today.month, today.day) < (born.month, born.day))
    except Exception:
        return None


def client_ip():
    fwd = request.headers.get("X-Forwarded-For", "")
    if fwd:
        return fwd.split(",")[0].strip()
    return request.remote_addr or "unknown"


def is_locked_out(db, username, ip):
    """Returns minutes remaining if locked out, else None."""
    row = db.execute(
        """SELECT COUNT(*) AS n, MAX(created_at) AS last_at FROM login_attempts
           WHERE success = 0 AND (username = ? OR ip = ?)
           AND created_at >= datetime('now', ?)""",
        (username, ip, f"-{LOGIN_LOCKOUT_MINUTES} minutes"),
    ).fetchone()
    if row and row["n"] >= LOGIN_MAX_ATTEMPTS:
        return LOGIN_LOCKOUT_MINUTES
    return None


def record_login_attempt(db, username, ip, success):
    db.execute(
        "INSERT INTO login_attempts (username, ip, success) VALUES (?, ?, ?)",
        (username, ip, 1 if success else 0),
    )
    db.commit()


# ---------------------------------------------------------------------------
# Simplified horoscope compatibility (Ashtakoot-style Guna Milan)
#
# NOTE: This checks 4 of the 8 traditional Ashtakoot factors (Varna, Gana,
# Nadi, Bhakoot) plus a Manglik check, using commonly published reference
# tables. It is meant as a quick screening aid for office staff, NOT a
# substitute for a full 36-guna reading from a qualified astrologer/pandit
# before finalizing any match. The UI must always show this disclaimer.
# ---------------------------------------------------------------------------
def _norm(s):
    return re.sub(r"[^a-z]", "", (s or "").lower())


RASHI_CANON = ["Mesha", "Vrishabha", "Mithuna", "Karka", "Simha", "Kanya",
               "Tula", "Vrischika", "Dhanu", "Makara", "Kumbha", "Meena"]

RASHI_ALIASES = {
    "mesha": "Mesha", "mesh": "Mesha", "aries": "Mesha",
    "vrishabha": "Vrishabha", "vrishabh": "Vrishabha", "vrushabh": "Vrishabha", "taurus": "Vrishabha",
    "mithuna": "Mithuna", "mithun": "Mithuna", "gemini": "Mithuna",
    "karka": "Karka", "kark": "Karka", "karkata": "Karka", "cancer": "Karka",
    "simha": "Simha", "sinh": "Simha", "leo": "Simha",
    "kanya": "Kanya", "virgo": "Kanya",
    "tula": "Tula", "libra": "Tula",
    "vrischika": "Vrischika", "vrishchik": "Vrischika", "vrischik": "Vrischika",
    "vrishchika": "Vrischika", "scorpio": "Vrischika",
    "dhanu": "Dhanu", "dhanus": "Dhanu", "sagittarius": "Dhanu",
    "makara": "Makara", "makar": "Makara", "capricorn": "Makara",
    "kumbha": "Kumbha", "aquarius": "Kumbha",
    "meena": "Meena", "meen": "Meena", "pisces": "Meena",
}

RASHI_VARNA = {
    "Karka": "Brahmin", "Vrischika": "Brahmin", "Meena": "Brahmin",
    "Mesha": "Kshatriya", "Simha": "Kshatriya", "Dhanu": "Kshatriya",
    "Vrishabha": "Vaishya", "Kanya": "Vaishya", "Makara": "Vaishya",
    "Mithuna": "Shudra", "Tula": "Shudra", "Kumbha": "Shudra",
}
VARNA_RANK = {"Brahmin": 4, "Kshatriya": 3, "Vaishya": 2, "Shudra": 1}

# (canonical name, gana, nadi)
NAKSHATRA_CANON = [
    ("Ashwini", "Deva", "Aadi"), ("Bharani", "Manushya", "Madhya"),
    ("Krittika", "Rakshasa", "Antya"), ("Rohini", "Manushya", "Madhya"),
    ("Mrigashira", "Deva", "Antya"), ("Ardra", "Manushya", "Aadi"),
    ("Punarvasu", "Deva", "Madhya"), ("Pushya", "Deva", "Antya"),
    ("Ashlesha", "Rakshasa", "Aadi"), ("Magha", "Rakshasa", "Madhya"),
    ("Purva Phalguni", "Manushya", "Antya"), ("Uttara Phalguni", "Manushya", "Aadi"),
    ("Hasta", "Deva", "Madhya"), ("Chitra", "Rakshasa", "Antya"),
    ("Swati", "Deva", "Aadi"), ("Vishakha", "Rakshasa", "Madhya"),
    ("Anuradha", "Deva", "Antya"), ("Jyeshtha", "Rakshasa", "Aadi"),
    ("Mula", "Rakshasa", "Madhya"), ("Purva Ashadha", "Manushya", "Antya"),
    ("Uttara Ashadha", "Manushya", "Aadi"), ("Shravana", "Deva", "Madhya"),
    ("Dhanishta", "Rakshasa", "Antya"), ("Shatabhisha", "Rakshasa", "Aadi"),
    ("Purva Bhadrapada", "Manushya", "Madhya"), ("Uttara Bhadrapada", "Manushya", "Antya"),
    ("Revati", "Deva", "Aadi"),
]
NAKSHATRA_ALIASES = {
    "ashwini": "Ashwini", "aswini": "Ashwini",
    "bharani": "Bharani",
    "krittika": "Krittika", "kartika": "Krittika", "kritika": "Krittika", "krithika": "Krittika",
    "rohini": "Rohini",
    "mrigashira": "Mrigashira", "mrigasira": "Mrigashira", "mrigashirsha": "Mrigashira", "mrugashira": "Mrigashira",
    "ardra": "Ardra", "arudra": "Ardra",
    "punarvasu": "Punarvasu",
    "pushya": "Pushya", "pushyami": "Pushya",
    "ashlesha": "Ashlesha", "aslesha": "Ashlesha",
    "magha": "Magha",
    "purvaphalguni": "Purva Phalguni", "pphalguni": "Purva Phalguni", "purvaphalguna": "Purva Phalguni",
    "uttaraphalguni": "Uttara Phalguni", "uphalguni": "Uttara Phalguni",
    "hasta": "Hasta",
    "chitra": "Chitra", "chitta": "Chitra",
    "swati": "Swati", "swathi": "Swati",
    "vishakha": "Vishakha", "visakha": "Vishakha", "vishaka": "Vishakha",
    "anuradha": "Anuradha",
    "jyeshtha": "Jyeshtha", "jyeshta": "Jyeshtha", "jyestha": "Jyeshtha",
    "mula": "Mula", "moola": "Mula",
    "purvaashadha": "Purva Ashadha", "pashadha": "Purva Ashadha", "purvashada": "Purva Ashadha",
    "uttaraashadha": "Uttara Ashadha", "uashadha": "Uttara Ashadha", "uttarashada": "Uttara Ashadha",
    "shravana": "Shravana", "sravana": "Shravana",
    "dhanishta": "Dhanishta", "dhanishtha": "Dhanishta", "dhanista": "Dhanishta",
    "shatabhisha": "Shatabhisha", "satabhisha": "Shatabhisha", "shatabhishak": "Shatabhisha",
    "purvabhadrapada": "Purva Bhadrapada", "pbhadrapada": "Purva Bhadrapada",
    "uttarabhadrapada": "Uttara Bhadrapada", "ubhadrapada": "Uttara Bhadrapada",
    "revati": "Revati",
}
NAKSHATRA_INFO = {name: (gana, nadi) for name, gana, nadi in NAKSHATRA_CANON}

GANA_SCORE = {
    ("Deva", "Deva"): 6, ("Manushya", "Manushya"): 6, ("Rakshasa", "Rakshasa"): 6,
    ("Deva", "Manushya"): 5, ("Manushya", "Deva"): 6,
    ("Deva", "Rakshasa"): 1, ("Rakshasa", "Deva"): 0,
    ("Manushya", "Rakshasa"): 0, ("Rakshasa", "Manushya"): 0,
}


def resolve_rashi(text):
    return RASHI_ALIASES.get(_norm(text))


def resolve_nakshatra(text):
    return NAKSHATRA_ALIASES.get(_norm(text))


def compute_compatibility(bride, groom):
    """Returns a dict with a simplified 4-factor Ashtakoot-style score.
    bride/groom are profile dicts with 'rashi', 'nakshatra', 'manglik'."""
    b_rashi = resolve_rashi(bride.get("rashi"))
    g_rashi = resolve_rashi(groom.get("rashi"))
    b_nak = resolve_nakshatra(bride.get("nakshatra"))
    g_nak = resolve_nakshatra(groom.get("nakshatra"))

    missing = []
    if not b_rashi or not g_rashi:
        missing.append("Rashi")
    if not b_nak or not g_nak:
        missing.append("Nakshatra")
    if missing:
        return {
            "available": False,
            "reason": f"Could not recognize {' and '.join(missing)} for one or both profiles. "
                      f"Please check spelling of these fields on the profile(s).",
        }

    factors = []

    # Varna (max 1)
    b_varna = RASHI_VARNA.get(b_rashi)
    g_varna = RASHI_VARNA.get(g_rashi)
    varna_score = 1 if (b_varna and g_varna and VARNA_RANK[g_varna] >= VARNA_RANK[b_varna]) else 0
    factors.append({"name": "Varna", "max": 1, "score": varna_score,
                     "note": f"{b_varna or '—'} (bride) vs {g_varna or '—'} (groom)"})

    # Gana (max 6)
    b_gana, b_nadi = NAKSHATRA_INFO[b_nak]
    g_gana, g_nadi = NAKSHATRA_INFO[g_nak]
    gana_score = GANA_SCORE.get((b_gana, g_gana), 3)
    factors.append({"name": "Gana", "max": 6, "score": gana_score,
                     "note": f"{b_gana} (bride) vs {g_gana} (groom)"})

    # Nadi (max 8) — same Nadi is traditionally considered inauspicious (Nadi Dosha)
    nadi_score = 0 if b_nadi == g_nadi else 8
    factors.append({"name": "Nadi", "max": 8, "score": nadi_score,
                     "note": f"{b_nadi} (bride) vs {g_nadi} (groom)" + (" — Nadi Dosha" if nadi_score == 0 else "")})

    # Bhakoot (max 7) — based on rashi distance in both directions
    bi, gi = RASHI_CANON.index(b_rashi), RASHI_CANON.index(g_rashi)
    fwd = ((gi - bi) % 12) + 1
    bwd = ((bi - gi) % 12) + 1
    dosha_counts = {2, 12, 5, 9, 6, 8}
    bhakoot_score = 0 if (fwd in dosha_counts or bwd in dosha_counts) else 7
    factors.append({"name": "Bhakoot", "max": 7, "score": bhakoot_score,
                     "note": f"{b_rashi} (bride) vs {g_rashi} (groom)" + (" — Bhakoot Dosha" if bhakoot_score == 0 else "")})

    total = sum(f["score"] for f in factors)
    max_total = sum(f["max"] for f in factors)

    manglik_b = (bride.get("manglik") or "").strip().lower()
    manglik_g = (groom.get("manglik") or "").strip().lower()
    manglik_note = "Manglik status not specified for one or both profiles."
    manglik_ok = None
    if manglik_b and manglik_g:
        manglik_ok = (manglik_b == manglik_g) or "partial" in (manglik_b, manglik_g)
        manglik_note = (
            f"Bride: {bride.get('manglik')}, Groom: {groom.get('manglik')} — "
            + ("compatible" if manglik_ok else "mismatch, traditionally a concern")
        )

    return {
        "available": True,
        "factors": factors,
        "total": total,
        "max_total": max_total,
        "percentage": round((total / max_total) * 100) if max_total else 0,
        "manglik_ok": manglik_ok,
        "manglik_note": manglik_note,
        "disclaimer": ("This is a simplified check covering 4 of the 8 traditional Ashtakoot "
                       "factors (Varna, Gana, Nadi, Bhakoot) plus a Manglik check. It is a quick "
                       "screening aid only — please consult a qualified astrologer/pandit for a "
                       "complete 36-guna Milan before finalizing any match."),
    }


# ---------------------------------------------------------------------------
# Static page routes (all pages are plain html files loose in this folder)
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "landing.html")


@app.route("/landing.html")
def landing_page():
    return send_from_directory(BASE_DIR, "landing.html")


@app.route("/register.html")
def register_page():
    return send_from_directory(BASE_DIR, "register.html")


@app.route("/signin.html")
def signin_page():
    return send_from_directory(BASE_DIR, "signin.html")


@app.route("/login.html")
def user_login_page():
    return send_from_directory(BASE_DIR, "login.html")


@app.route("/dashboard.html")
@user_login_required
def user_dashboard_page():
    return send_from_directory(BASE_DIR, "dashboard.html")


@app.route("/admin.html")
@login_required
def admin_page():
    return send_from_directory(BASE_DIR, "admin.html")


@app.route("/view.html")
@login_required
def view_page():
    return send_from_directory(BASE_DIR, "view.html")


@app.route("/matches.html")
@login_required
def matches_page():
    return send_from_directory(BASE_DIR, "matches.html")


@app.route("/reports.html")
@login_required
def reports_page():
    return send_from_directory(BASE_DIR, "reports.html")


@app.route("/details.html")
def details_page():
    # Public read-only page (also the QR destination). No admin functions here.
    return send_from_directory(BASE_DIR, "details.html")


@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)


# ---------------------------------------------------------------------------
# Auth API
# ---------------------------------------------------------------------------
@app.route("/signin", methods=["POST"])
def signin():
    data = request.get_json(silent=True) or request.form
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    ip = client_ip()
    db = get_db()

    lockout = is_locked_out(db, username, ip)
    if lockout:
        return jsonify({
            "success": False,
            "message": f"Too many failed attempts. Please try again in about {lockout} minutes."
        }), 429

    row = db.execute("SELECT * FROM admins WHERE username = ?", (username,)).fetchone()
    if row and check_password_hash(row["password_hash"], password):
        record_login_attempt(db, username, ip, True)
        session["admin_id"] = row["id"]
        session["admin_username"] = row["username"]
        return jsonify({"success": True, "redirect": "/admin.html"})

    record_login_attempt(db, username, ip, False)
    return jsonify({"success": False, "message": "Invalid username or password"}), 401


@app.route("/signout", methods=["POST"])
def signout():
    session.clear()
    return jsonify({"success": True})


@app.route("/api/me")
def api_me():
    if session.get("admin_id"):
        return jsonify({"authenticated": True, "username": session.get("admin_username")})
    return jsonify({"authenticated": False})


@app.route("/change-password", methods=["POST"])
@login_required
def change_password():
    data = request.get_json(silent=True) or request.form
    current = data.get("current_password") or ""
    new = data.get("new_password") or ""
    db = get_db()
    row = db.execute("SELECT * FROM admins WHERE id = ?", (session["admin_id"],)).fetchone()
    if not row or not check_password_hash(row["password_hash"], current):
        return jsonify({"success": False, "message": "Current password is incorrect"}), 400
    if len(new) < 6:
        return jsonify({"success": False, "message": "New password must be at least 6 characters"}), 400
    db.execute("UPDATE admins SET password_hash = ? WHERE id = ?", (generate_password_hash(new), row["id"]))
    db.commit()
    return jsonify({"success": True})


@app.route("/create-admin", methods=["POST"])
@login_required
def create_admin():
    data = request.get_json(silent=True) or request.form
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    if not username or len(password) < 6:
        return jsonify({"success": False, "message": "Username required & password min 6 chars"}), 400
    db = get_db()
    try:
        db.execute("INSERT INTO admins (username, password_hash) VALUES (?, ?)",
                   (username, generate_password_hash(password)))
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({"success": False, "message": "Username already exists"}), 400
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# User (registrant) authentication & self-service dashboard
# ---------------------------------------------------------------------------
@app.route("/user-login", methods=["POST"])
def user_login():
    data = request.get_json(silent=True) or request.form
    identifier = (data.get("identifier") or "").strip()
    password = data.get("password") or ""
    ip = client_ip()
    db = get_db()

    lockout = is_locked_out(db, f"user:{identifier}", ip)
    if lockout:
        return jsonify({"success": False, "message": f"Too many failed attempts. Try again in about {lockout} minutes."}), 429

    row = db.execute(
        "SELECT * FROM profiles WHERE (email = ? OR phone = ?) AND password_hash IS NOT NULL",
        (identifier, identifier),
    ).fetchone()
    if row and check_password_hash(row["password_hash"], password):
        record_login_attempt(db, f"user:{identifier}", ip, True)
        session["user_profile_id"] = row["id"]
        return jsonify({"success": True, "redirect": "/dashboard.html"})

    record_login_attempt(db, f"user:{identifier}", ip, False)
    return jsonify({"success": False, "message": "Invalid email/phone or password"}), 401


@app.route("/user-logout", methods=["POST"])
def user_logout():
    session.pop("user_profile_id", None)
    return jsonify({"success": True})


@app.route("/api/my-profile")
@user_login_required
def api_my_profile():
    db = get_db()
    row = db.execute("SELECT * FROM profiles WHERE id = ?", (session["user_profile_id"],)).fetchone()
    if not row:
        session.pop("user_profile_id", None)
        return jsonify({"success": False, "message": "Profile not found"}), 404
    profile = attach_photos(db, row_to_dict(row))
    profile.pop("password_hash", None)
    return jsonify({"success": True, "profile": profile})


@app.route("/api/my-profile", methods=["PUT"])
@user_login_required
def api_my_profile_update():
    data = request.get_json(silent=True) or {}
    db = get_db()
    pid = session["user_profile_id"]
    sets, vals = [], []
    for k, v in data.items():
        if k in PROFILE_FIELDS:
            sets.append(f"{k} = ?")
            vals.append(v)
    if not sets:
        return jsonify({"success": False, "message": "Nothing to update"}), 400
    # any self-edit sends the profile back for admin re-approval
    sets.append("status = ?")
    vals.append("Pending")
    sets.append("updated_at = ?")
    vals.append(datetime.utcnow().isoformat())
    vals.append(pid)
    db.execute(f"UPDATE profiles SET {', '.join(sets)} WHERE id = ?", vals)
    db.commit()
    return jsonify({"success": True, "message": "Saved. Your profile will be reviewed by the bureau before going live again."})


@app.route("/api/my-profile/password", methods=["POST"])
@user_login_required
def api_my_profile_password():
    data = request.get_json(silent=True) or {}
    current = data.get("current_password") or ""
    new = data.get("new_password") or ""
    db = get_db()
    row = db.execute("SELECT * FROM profiles WHERE id = ?", (session["user_profile_id"],)).fetchone()
    if not row or not row["password_hash"] or not check_password_hash(row["password_hash"], current):
        return jsonify({"success": False, "message": "Current password is incorrect"}), 400
    if len(new) < 6:
        return jsonify({"success": False, "message": "New password must be at least 6 characters"}), 400
    db.execute("UPDATE profiles SET password_hash = ? WHERE id = ?", (generate_password_hash(new), row["id"]))
    db.commit()
    return jsonify({"success": True})


@app.route("/api/my-profile/photos", methods=["POST"])
@user_login_required
def api_my_profile_add_photos():
    pid = session["user_profile_id"]
    db = get_db()
    files = request.files.getlist("photos")
    if not files:
        return jsonify({"success": False, "message": "Please choose at least one photo."}), 400
    existing_count = db.execute("SELECT COUNT(*) FROM photos WHERE profile_id = ?", (pid,)).fetchone()[0]
    added = []
    for i, fs in enumerate(files):
        if fs and fs.filename and allowed_ext(fs.filename, ALLOWED_IMAGE_EXT):
            saved = save_upload(fs, f"photo_{pid}")
            is_primary = 1 if existing_count == 0 and i == 0 else 0
            db.execute(
                "INSERT INTO photos (profile_id, filepath, is_primary, sort_order) VALUES (?, ?, ?, ?)",
                (pid, saved, is_primary, existing_count + i),
            )
            added.append(f"/uploads/{saved}")
    db.commit()
    return jsonify({"success": True, "added": added})


def _photo_owner_check(db, photo_id):
    """Returns the owning profile_id if the current session (admin or matching user) may manage this photo."""
    row = db.execute("SELECT * FROM photos WHERE id = ?", (photo_id,)).fetchone()
    if not row:
        return None, None
    if session.get("admin_id"):
        return row, row["profile_id"]
    if session.get("user_profile_id") == row["profile_id"]:
        return row, row["profile_id"]
    return row, None


@app.route("/api/my-photo/<int:photo_id>", methods=["DELETE"])
@user_login_required
def api_my_photo_delete(photo_id):
    db = get_db()
    row, owner = _photo_owner_check(db, photo_id)
    if not row or owner != session["user_profile_id"]:
        abort(404)
    was_primary = row["is_primary"]
    db.execute("DELETE FROM photos WHERE id = ?", (photo_id,))
    if was_primary:
        nxt = db.execute("SELECT id FROM photos WHERE profile_id = ? ORDER BY sort_order ASC LIMIT 1", (owner,)).fetchone()
        if nxt:
            db.execute("UPDATE photos SET is_primary = 1 WHERE id = ?", (nxt["id"],))
    db.commit()
    return jsonify({"success": True})


@app.route("/api/my-photo/<int:photo_id>/primary", methods=["POST"])
@user_login_required
def api_my_photo_primary(photo_id):
    db = get_db()
    row, owner = _photo_owner_check(db, photo_id)
    if not row or owner != session["user_profile_id"]:
        abort(404)
    db.execute("UPDATE photos SET is_primary = 0 WHERE profile_id = ?", (owner,))
    db.execute("UPDATE photos SET is_primary = 1 WHERE id = ?", (photo_id,))
    db.commit()
    return jsonify({"success": True})


@app.route("/api/photos/reorder", methods=["POST"])
def api_photos_reorder():
    """Shared reorder endpoint for both admin (view.html) and user (dashboard.html) galleries."""
    data = request.get_json(silent=True) or {}
    profile_id = data.get("profile_id")
    order = data.get("order") or []
    if not profile_id or not order:
        return jsonify({"success": False, "message": "Missing profile_id or order"}), 400

    is_admin = bool(session.get("admin_id"))
    is_owner = session.get("user_profile_id") == int(profile_id)
    if not is_admin and not is_owner:
        return jsonify({"success": False, "message": "Not authorized"}), 401

    db = get_db()
    for idx, photo_id in enumerate(order):
        db.execute(
            "UPDATE photos SET sort_order = ? WHERE id = ? AND profile_id = ?",
            (idx, photo_id, profile_id),
        )
    db.commit()
    return jsonify({"success": True})



# ---------------------------------------------------------------------------
@app.route("/register", methods=["POST"])
def register():
    form = request.form
    if not form.get("full_name") or not form.get("phone") or not form.get("email"):
        return jsonify({"success": False, "message": "Please fill all required fields."}), 400
    if not form.get("password") or len(form.get("password")) < 6:
        return jsonify({"success": False, "message": "Please set a password (min 6 characters) so you can log in later to edit your profile."}), 400
    if not re.match(r"^[6-9]\d{9}$", form.get("phone", "").strip()):
        return jsonify({"success": False, "message": "Please enter a valid 10-digit phone number."}), 400
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", form.get("email", "").strip()):
        return jsonify({"success": False, "message": "Please enter a valid email address."}), 400

    db = get_db()
    cols, vals, marks = [], [], []
    for field in PROFILE_FIELDS:
        cols.append(field)
        v = form.get(field, "")
        if field == "age" and not v:
            v = calc_age(form.get("dob", "")) or None
        vals.append(v)
        marks.append("?")

    # password (optional, hashed if provided)
    pw = form.get("password") or ""
    cols.append("password_hash")
    vals.append(generate_password_hash(pw) if pw else None)
    marks.append("?")

    # file uploads (single)
    for f in FILE_FIELDS:
        saved = save_upload(request.files.get(f), f)
        col = f + "_path" if f != "kundli_pdf" else "kundli_pdf_path"
        cols.append(col)
        vals.append(saved)
        marks.append("?")

    token = uuid.uuid4().hex
    cols.append("public_token")
    vals.append(token)
    marks.append("?")

    cols.append("status")
    vals.append("Pending")
    marks.append("?")

    sql = f"INSERT INTO profiles ({', '.join(cols)}) VALUES ({', '.join(marks)})"
    cur = db.execute(sql, vals)
    profile_id = cur.lastrowid

    db.commit()

    # profile photos (multiple allowed)
    files = request.files.getlist("photos")
    for i, fs in enumerate(files):
        if fs and fs.filename and allowed_ext(fs.filename, ALLOWED_IMAGE_EXT):
            saved = save_upload(fs, f"photo_{profile_id}")
            db.execute(
                "INSERT INTO photos (profile_id, filepath, is_primary, sort_order) VALUES (?, ?, ?, ?)",
                (profile_id, saved, 1 if i == 0 else 0, i),
            )
    db.commit()

    return jsonify({"success": True, "profile_id": profile_id})


# ---------------------------------------------------------------------------
# Profiles API (admin + public detail)
# ---------------------------------------------------------------------------
def row_to_dict(row):
    d = dict(row)
    return d


def compute_completeness(profile):
    filled = sum(1 for f in PROFILE_FIELDS if (profile.get(f) not in (None, "", "None")))
    has_photo = 1 if profile.get("photos") else 0
    total_slots = len(PROFILE_FIELDS) + 1
    filled_slots = filled + has_photo
    return round((filled_slots / total_slots) * 100)


def attach_photos(db, profile):
    rows = db.execute(
        "SELECT id, filepath, is_primary FROM photos WHERE profile_id = ? ORDER BY sort_order ASC",
        (profile["id"],),
    ).fetchall()
    profile["photos"] = [f"/uploads/{r['filepath']}" for r in rows]
    profile["photo_list"] = [{"id": r["id"], "url": f"/uploads/{r['filepath']}", "is_primary": bool(r["is_primary"])} for r in rows]
    primary_row = next((r for r in rows if r["is_primary"]), rows[0] if rows else None)
    profile["primary_photo"] = f"/uploads/{primary_row['filepath']}" if primary_row else None
    for f in ["kundli_pdf_path", "aadhar_photo_path", "passport_photo_path"]:
        if profile.get(f):
            profile[f] = f"/uploads/{profile[f]}"
    profile["completeness"] = compute_completeness(profile)
    return profile


@app.route("/api/profiles")
@login_required
def api_profiles():
    db = get_db()
    q = "SELECT * FROM profiles WHERE 1=1"
    args = []

    def add_eq(param, col):
        nonlocal q
        val = request.args.get(param)
        if val:
            q_local = f" AND {col} = ?"
            return q_local, val
        return "", None

    filters = {
        "gender": "gender", "profile_for": "profile_for", "marital_status": "marital_status",
        "current_city": "current_city", "native_place": "native_place",
        "sindhi_caste": "sindhi_caste", "sub_caste": "sub_caste",
        "manglik": "manglik", "food_preference": "food_preference",
        "smoking": "smoking", "drinking": "drinking", "blood_group": "blood_group",
        "complexion": "complexion", "body_type": "body_type", "status": "status",
        "qualification": "qualification", "occupation": "occupation",
    }
    for param, col in filters.items():
        val = request.args.get(param)
        if val:
            q += f" AND {col} = ?"
            args.append(val)

    if not request.args.get("status"):
        q += " AND status NOT IN ('Pending', 'Rejected')"

    age_min = request.args.get("age_min")
    age_max = request.args.get("age_max")
    if age_min:
        q += " AND age >= ?"
        args.append(age_min)
    if age_max:
        q += " AND age <= ?"
        args.append(age_max)

    birth_year_min = request.args.get("birth_year_min")
    birth_year_max = request.args.get("birth_year_max")
    if birth_year_min:
        q += " AND CAST(substr(dob, 1, 4) AS INTEGER) >= ?"
        args.append(birth_year_min)
    if birth_year_max:
        q += " AND CAST(substr(dob, 1, 4) AS INTEGER) <= ?"
        args.append(birth_year_max)

    search = request.args.get("search")
    if search:
        q += " AND (full_name LIKE ? OR phone LIKE ? OR email LIKE ? OR occupation LIKE ?)"
        like = f"%{search}%"
        args.extend([like, like, like, like])

    q += " ORDER BY created_at DESC"
    rows = db.execute(q, args).fetchall()
    profiles = [attach_photos(db, row_to_dict(r)) for r in rows]
    return jsonify({"success": True, "count": len(profiles), "profiles": profiles})


@app.route("/api/filter-options")
@login_required
def api_filter_options():
    db = get_db()
    cols = ["current_city", "native_place", "sindhi_caste", "sub_caste", "marital_status",
            "qualification", "occupation", "blood_group", "complexion", "body_type",
            "food_preference", "manglik"]
    options = {}
    for c in cols:
        rows = db.execute(f"SELECT DISTINCT {c} FROM profiles WHERE {c} IS NOT NULL AND {c} != '' ORDER BY {c}").fetchall()
        options[c] = [r[0] for r in rows]
    return jsonify({"success": True, "options": options})


@app.route("/api/profile/<int:profile_id>")
@login_required
def api_profile_detail(profile_id):
    db = get_db()
    row = db.execute("SELECT * FROM profiles WHERE id = ?", (profile_id,)).fetchone()
    if not row:
        abort(404)
    return jsonify({"success": True, "profile": attach_photos(db, row_to_dict(row))})


@app.route("/api/public/profile")
def api_public_profile():
    """Used by details.html (QR destination) - read-only, token-gated, no sensitive admin data."""
    token = request.args.get("token")
    pid = request.args.get("id")
    db = get_db()
    if token:
        row = db.execute("SELECT * FROM profiles WHERE public_token = ?", (token,)).fetchone()
    elif pid and session.get("admin_id"):
        row = db.execute("SELECT * FROM profiles WHERE id = ?", (pid,)).fetchone()
    else:
        return jsonify({"success": False, "message": "Missing token"}), 400
    if not row:
        abort(404)
    profile = attach_photos(db, row_to_dict(row))
    # strip sensitive verification data from the public view
    for f in ["aadhar_number", "aadhar_photo_path", "passport_number", "passport_photo_path",
              "password_hash", "public_token"]:
        profile.pop(f, None)
    return jsonify({"success": True, "profile": profile})


@app.route("/api/profile/<int:profile_id>", methods=["PUT"])
@login_required
def api_profile_update(profile_id):
    data = request.get_json(silent=True) or {}
    db = get_db()
    allowed_cols = set(PROFILE_FIELDS) | {"status", "admin_notes"}
    sets, vals = [], []
    for k, v in data.items():
        if k in allowed_cols:
            sets.append(f"{k} = ?")
            vals.append(v)
    if not sets:
        return jsonify({"success": False, "message": "Nothing to update"}), 400
    sets.append("updated_at = ?")
    vals.append(datetime.utcnow().isoformat())
    vals.append(profile_id)
    db.execute(f"UPDATE profiles SET {', '.join(sets)} WHERE id = ?", vals)
    db.commit()
    return jsonify({"success": True})


@app.route("/api/profile/<int:profile_id>/approve", methods=["POST"])
@login_required
def api_profile_approve(profile_id):
    db = get_db()
    row = db.execute("SELECT id FROM profiles WHERE id = ?", (profile_id,)).fetchone()
    if not row:
        abort(404)
    db.execute("UPDATE profiles SET status = 'Active', updated_at = ? WHERE id = ?",
               (datetime.utcnow().isoformat(), profile_id))
    db.commit()
    return jsonify({"success": True})


@app.route("/api/profile/<int:profile_id>/reject", methods=["POST"])
@login_required
def api_profile_reject(profile_id):
    db = get_db()
    row = db.execute("SELECT id FROM profiles WHERE id = ?", (profile_id,)).fetchone()
    if not row:
        abort(404)
    db.execute("UPDATE profiles SET status = 'Rejected', updated_at = ? WHERE id = ?",
               (datetime.utcnow().isoformat(), profile_id))
    db.commit()
    return jsonify({"success": True})


@app.route("/api/profile/<int:profile_id>", methods=["DELETE"])
@login_required
def api_profile_delete(profile_id):
    db = get_db()
    db.execute("DELETE FROM profiles WHERE id = ?", (profile_id,))
    db.commit()
    return jsonify({"success": True})


@app.route("/api/profile/<int:profile_id>/photos", methods=["POST"])
@login_required
def api_profile_add_photos(profile_id):
    db = get_db()
    row = db.execute("SELECT id FROM profiles WHERE id = ?", (profile_id,)).fetchone()
    if not row:
        abort(404)
    files = request.files.getlist("photos")
    if not files:
        return jsonify({"success": False, "message": "Please choose at least one photo."}), 400

    existing_count = db.execute("SELECT COUNT(*) FROM photos WHERE profile_id = ?", (profile_id,)).fetchone()[0]
    added = []
    for i, fs in enumerate(files):
        if fs and fs.filename and allowed_ext(fs.filename, ALLOWED_IMAGE_EXT):
            saved = save_upload(fs, f"photo_{profile_id}")
            is_primary = 1 if existing_count == 0 and i == 0 else 0
            db.execute(
                "INSERT INTO photos (profile_id, filepath, is_primary, sort_order) VALUES (?, ?, ?, ?)",
                (profile_id, saved, is_primary, existing_count + i),
            )
            added.append(f"/uploads/{saved}")
    db.commit()
    return jsonify({"success": True, "added": added})


@app.route("/api/photo/<int:photo_id>", methods=["DELETE"])
@login_required
def api_delete_photo(photo_id):
    db = get_db()
    row = db.execute("SELECT * FROM photos WHERE id = ?", (photo_id,)).fetchone()
    if not row:
        abort(404)
    profile_id = row["profile_id"]
    was_primary = row["is_primary"]
    db.execute("DELETE FROM photos WHERE id = ?", (photo_id,))
    if was_primary:
        nxt = db.execute(
            "SELECT id FROM photos WHERE profile_id = ? ORDER BY sort_order ASC LIMIT 1", (profile_id,)
        ).fetchone()
        if nxt:
            db.execute("UPDATE photos SET is_primary = 1 WHERE id = ?", (nxt["id"],))
    db.commit()
    return jsonify({"success": True})


@app.route("/api/photo/<int:photo_id>/primary", methods=["POST"])
@login_required
def api_set_primary_photo(photo_id):
    db = get_db()
    row = db.execute("SELECT * FROM photos WHERE id = ?", (photo_id,)).fetchone()
    if not row:
        abort(404)
    db.execute("UPDATE photos SET is_primary = 0 WHERE profile_id = ?", (row["profile_id"],))
    db.execute("UPDATE photos SET is_primary = 1 WHERE id = ?", (photo_id,))
    db.commit()
    return jsonify({"success": True})


@app.route("/api/profile/<int:profile_id>/qr.png")
def api_profile_qr(profile_id):
    import qrcode
    db = get_db()
    row = db.execute("SELECT public_token FROM profiles WHERE id = ?", (profile_id,)).fetchone()
    if not row:
        abort(404)
    base = request.url_root.rstrip("/")
    target = f"{base}/details.html?token={row['public_token']}"
    img = qrcode.make(target)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")


# ---------------------------------------------------------------------------
# Matching / Shortlisting
# ---------------------------------------------------------------------------
def match_row_to_dict(db, row):
    d = dict(row)
    a = db.execute("SELECT * FROM profiles WHERE id = ?", (d["profile_a_id"],)).fetchone()
    b = db.execute("SELECT * FROM profiles WHERE id = ?", (d["profile_b_id"],)).fetchone()
    d["profile_a"] = attach_photos(db, row_to_dict(a)) if a else None
    d["profile_b"] = attach_photos(db, row_to_dict(b)) if b else None
    return d


@app.route("/api/matches")
@login_required
def api_matches_list():
    db = get_db()
    status = request.args.get("status")
    q = "SELECT * FROM matches WHERE 1=1"
    args = []
    if status:
        q += " AND status = ?"
        args.append(status)
    q += " ORDER BY created_at DESC"
    rows = db.execute(q, args).fetchall()
    matches = [match_row_to_dict(db, r) for r in rows]
    return jsonify({"success": True, "count": len(matches), "matches": matches})


@app.route("/api/matches", methods=["POST"])
@login_required
def api_matches_create():
    data = request.get_json(silent=True) or {}
    a_id = data.get("profile_a_id")
    b_id = data.get("profile_b_id")
    notes = data.get("notes", "")
    if not a_id or not b_id or int(a_id) == int(b_id):
        return jsonify({"success": False, "message": "Please choose two different profiles."}), 400
    db = get_db()
    existing = db.execute(
        """SELECT id FROM matches WHERE
           (profile_a_id = ? AND profile_b_id = ?) OR (profile_a_id = ? AND profile_b_id = ?)""",
        (a_id, b_id, b_id, a_id),
    ).fetchone()
    if existing:
        return jsonify({"success": False, "message": "These two profiles are already shortlisted together."}), 400
    cur = db.execute(
        "INSERT INTO matches (profile_a_id, profile_b_id, notes) VALUES (?, ?, ?)",
        (a_id, b_id, notes),
    )
    db.commit()
    return jsonify({"success": True, "match_id": cur.lastrowid})


@app.route("/api/matches/<int:match_id>", methods=["PUT"])
@login_required
def api_matches_update(match_id):
    data = request.get_json(silent=True) or {}
    db = get_db()
    sets, vals = [], []
    if "status" in data:
        sets.append("status = ?")
        vals.append(data["status"])
    if "notes" in data:
        sets.append("notes = ?")
        vals.append(data["notes"])
    if not sets:
        return jsonify({"success": False, "message": "Nothing to update"}), 400
    sets.append("updated_at = ?")
    vals.append(datetime.utcnow().isoformat())
    vals.append(match_id)
    db.execute(f"UPDATE matches SET {', '.join(sets)} WHERE id = ?", vals)
    db.commit()
    return jsonify({"success": True})


@app.route("/api/matches/<int:match_id>", methods=["DELETE"])
@login_required
def api_matches_delete(match_id):
    db = get_db()
    db.execute("DELETE FROM matches WHERE id = ?", (match_id,))
    db.commit()
    return jsonify({"success": True})


@app.route("/api/profile/<int:profile_id>/suggested-matches")
@login_required
def api_suggested_matches(profile_id):
    db = get_db()
    anchor = db.execute("SELECT * FROM profiles WHERE id = ?", (profile_id,)).fetchone()
    if not anchor:
        abort(404)
    anchor = dict(anchor)
    opposite = "Groom" if anchor.get("profile_for") == "Bride" else "Bride"

    rows = db.execute(
        "SELECT * FROM profiles WHERE profile_for = ? AND id != ? AND status = 'Active'",
        (opposite, profile_id),
    ).fetchall()

    candidates = []
    anchor_age = anchor.get("age")
    for r in rows:
        c = dict(r)
        score = 0
        if anchor_age and c.get("age"):
            diff = abs(int(anchor_age) - int(c["age"]))
            if diff <= 5:
                score += max(0, 10 - diff)
        if anchor.get("sindhi_caste") and c.get("sindhi_caste") == anchor.get("sindhi_caste"):
            score += 6
        if anchor.get("current_city") and c.get("current_city") == anchor.get("current_city"):
            score += 4
        if anchor.get("marital_status") and c.get("marital_status") == anchor.get("marital_status"):
            score += 3
        if anchor.get("manglik") and c.get("manglik") == anchor.get("manglik"):
            score += 2
        c["match_score"] = score
        candidates.append(c)

    candidates.sort(key=lambda x: x["match_score"], reverse=True)
    top = [attach_photos(db, c) for c in candidates[:12]]
    return jsonify({"success": True, "candidates": top})


@app.route("/api/compatibility/<int:a_id>/<int:b_id>")
@login_required
def api_compatibility(a_id, b_id):
    db = get_db()
    a = db.execute("SELECT * FROM profiles WHERE id = ?", (a_id,)).fetchone()
    b = db.execute("SELECT * FROM profiles WHERE id = ?", (b_id,)).fetchone()
    if not a or not b:
        abort(404)
    a, b = dict(a), dict(b)
    if a.get("profile_for") == "Groom" and b.get("profile_for") == "Bride":
        bride, groom = b, a
    else:
        bride, groom = a, b
    result = compute_compatibility(bride, groom)
    return jsonify({"success": True, **result})


# ---------------------------------------------------------------------------
# CSV Import / Export
# ---------------------------------------------------------------------------
@app.route("/export/csv")
@login_required
def export_csv():
    db = get_db()
    rows = db.execute("SELECT * FROM profiles ORDER BY created_at DESC").fetchall()
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=PROFILE_FIELDS)
    writer.writeheader()
    for r in rows:
        d = dict(r)
        writer.writerow({k: d.get(k, "") for k in PROFILE_FIELDS})
    mem = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    fname = f"profiles_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(mem, mimetype="text/csv", as_attachment=True, download_name=fname)


@app.route("/import/csv", methods=["POST"])
@login_required
def import_csv():
    file = request.files.get("file")
    if not file or not file.filename.lower().endswith(".csv"):
        return jsonify({"success": False, "message": "Please upload a .csv file"}), 400
    stream = io.StringIO(file.stream.read().decode("utf-8-sig"))
    reader = csv.DictReader(stream)
    db = get_db()
    inserted = 0
    for row in reader:
        cols, vals, marks = [], [], []
        for field in PROFILE_FIELDS:
            cols.append(field)
            vals.append(row.get(field, ""))
            marks.append("?")
        cols.append("public_token")
        vals.append(uuid.uuid4().hex)
        marks.append("?")
        db.execute(f"INSERT INTO profiles ({', '.join(cols)}) VALUES ({', '.join(marks)})", vals)
        inserted += 1
    db.commit()
    return jsonify({"success": True, "inserted": inserted})


@app.route("/export/xlsx")
@login_required
def export_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    db = get_db()
    rows = db.execute("SELECT * FROM profiles ORDER BY created_at DESC").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Profiles"

    header_fill = PatternFill(start_color="1B3A63", end_color="1B3A63", fill_type="solid")
    header_font = Font(color="F6F0E4", bold=True)

    for col_idx, field in enumerate(PROFILE_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, r in enumerate(rows, start=2):
        d = dict(r)
        for col_idx, field in enumerate(PROFILE_FIELDS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=d.get(field, ""))

    for col_idx, field in enumerate(PROFILE_FIELDS, start=1):
        width = max(12, min(28, len(field) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"profiles_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=fname,
    )


@app.route("/import/xlsx", methods=["POST"])
@login_required
def import_xlsx():
    from openpyxl import load_workbook

    file = request.files.get("file")
    if not file or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"success": False, "message": "Please upload a .xlsx file"}), 400

    wb = load_workbook(file, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h else "" for h in next(rows_iter)]

    db = get_db()
    inserted = 0
    for raw_row in rows_iter:
        if raw_row is None or all(v in (None, "") for v in raw_row):
            continue
        row = dict(zip(header, raw_row))
        cols, vals, marks = [], [], []
        for field in PROFILE_FIELDS:
            cols.append(field)
            val = row.get(field, "")
            vals.append("" if val is None else val)
            marks.append("?")
        cols.append("public_token")
        vals.append(uuid.uuid4().hex)
        marks.append("?")
        db.execute(f"INSERT INTO profiles ({', '.join(cols)}) VALUES ({', '.join(marks)})", vals)
        inserted += 1
    db.commit()
    return jsonify({"success": True, "inserted": inserted})


# ---------------------------------------------------------------------------
# Backup / Restore
# ---------------------------------------------------------------------------
@app.route("/backup")
@login_required
def backup_db():
    fname = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    return send_file(DB_PATH, as_attachment=True, download_name=fname)


@app.route("/restore", methods=["POST"])
@login_required
def restore_db():
    file = request.files.get("file")
    if not file or not allowed_ext(file.filename, ALLOWED_DB_EXT):
        return jsonify({"success": False, "message": "Please upload a valid .db backup file"}), 400
    close_db()
    file.save(DB_PATH)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Biodata generation (PDF & PNG)
# ---------------------------------------------------------------------------
def biodata_lines(profile):
    def g(k, label=None):
        v = profile.get(k)
        return (label or k.replace("_", " ").title(), v if v not in (None, "") else "-")

    sections = [
        ("Basic Information", [
            g("full_name", "Full Name"), g("gender", "Gender"), g("dob", "Date of Birth"),
            g("age", "Age"), g("height", "Height"), g("complexion", "Complexion"),
            g("marital_status", "Marital Status"), g("father_name", "Father's Name"),
            g("mother_name", "Mother's Name"), g("sindhi_caste", "Sindhi Caste"),
            g("mother_tongue", "Mother Tongue"),
        ]),
        ("Education & Career", [
            g("qualification", "Qualification"), g("occupation", "Occupation"),
            g("company", "Company"), g("designation", "Designation"),
            g("income_yearly", "Annual Income"),
        ]),
        ("Contact & Address", [
            g("current_city", "Current City"), g("native_place", "Native Place"),
            g("phone", "Phone"), g("email", "Email"),
        ]),
        ("Astrological Details", [
            g("manglik", "Manglik"), g("rashi", "Rashi"), g("nakshatra", "Nakshatra"),
        ]),
        ("About", [g("about_yourself", "About")]),
    ]
    return sections


@app.route("/biodata/<int:profile_id>/pdf")
@login_required
def biodata_pdf(profile_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor
    from reportlab.pdfgen import canvas as pdfcanvas

    db = get_db()
    row = db.execute("SELECT * FROM profiles WHERE id = ?", (profile_id,)).fetchone()
    if not row:
        abort(404)
    profile = attach_photos(db, row_to_dict(row))

    buf = io.BytesIO()
    c = pdfcanvas.Canvas(buf, pagesize=A4)
    width, height = A4
    gold = HexColor("#c9972b")
    indigo = HexColor("#1b3a63")
    maroon = HexColor("#8b1e3f")
    ink = HexColor("#14161f")

    c.setFillColor(maroon)
    c.rect(0, height - 14 * mm, width, 14 * mm, fill=1, stroke=0)
    c.setFillColor(HexColor("#f6f0e4"))
    c.setFont("Helvetica-Bold", 16)
    c.drawString(18 * mm, height - 10 * mm, "Sindhi Samaj Pariwar Surat — Marriage Biodata")

    y = height - 26 * mm

    # photo
    photo_path = None
    if profile["photos"]:
        photo_path = os.path.join(UPLOAD_DIR, profile["photos"][0].split("/uploads/")[-1])
    if photo_path and os.path.exists(photo_path):
        try:
            c.drawImage(photo_path, width - 55 * mm, y - 45 * mm, width=40 * mm, height=48 * mm,
                        preserveAspectRatio=True, anchor='n')
        except Exception:
            pass

    c.setFillColor(indigo)
    c.setFont("Helvetica-Bold", 20)
    c.drawString(18 * mm, y, profile.get("full_name") or "Unnamed Profile")
    y -= 10 * mm

    for section_title, rows in biodata_lines(profile):
        if y < 25 * mm:
            c.showPage()
            y = height - 20 * mm
        c.setFillColor(gold)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(18 * mm, y, section_title)
        y -= 6 * mm
        c.setStrokeColor(gold)
        c.line(18 * mm, y + 2 * mm, width - 18 * mm, y + 2 * mm)
        c.setFont("Helvetica", 10)
        for label, val in rows:
            if y < 20 * mm:
                c.showPage()
                y = height - 20 * mm
            c.setFillColor(ink)
            text = f"{label}: {val}"
            for line in _wrap_text(text, 100):
                c.drawString(20 * mm, y, line)
                y -= 5.2 * mm
        y -= 3 * mm

    c.save()
    buf.seek(0)
    fname = f"biodata_{secure_filename(profile.get('full_name') or str(profile_id))}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=fname)


def _wrap_text(text, width):
    words = text.split(" ")
    lines, cur = [], ""
    for w in words:
        if len(cur) + len(w) + 1 <= width:
            cur = (cur + " " + w).strip()
        else:
            lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return lines


@app.route("/biodata/<int:profile_id>/png")
@login_required
def biodata_png(profile_id):
    from PIL import Image, ImageDraw, ImageFont

    db = get_db()
    row = db.execute("SELECT * FROM profiles WHERE id = ?", (profile_id,)).fetchone()
    if not row:
        abort(404)
    profile = attach_photos(db, row_to_dict(row))

    W, H = 1000, 1400
    img = Image.new("RGB", (W, H), "#fbf8f1")
    draw = ImageDraw.Draw(img)

    def font(size, bold=False):
        try:
            path = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold \
                else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
            return ImageFont.truetype(path, size)
        except Exception:
            return ImageFont.load_default()

    maroon = (139, 30, 63)
    indigo = (27, 58, 99)
    gold = (201, 151, 43)
    ink = (20, 22, 31)

    draw.rectangle([0, 0, W, 70], fill=maroon)
    draw.text((30, 20), "Sindhi Samaj Pariwar Surat — Marriage Biodata", font=font(26, True), fill="white")

    y = 100
    photo_path = None
    if profile["photos"]:
        photo_path = os.path.join(UPLOAD_DIR, profile["photos"][0].split("/uploads/")[-1])
    if photo_path and os.path.exists(photo_path):
        try:
            ph = Image.open(photo_path).convert("RGB")
            ph.thumbnail((260, 320))
            img.paste(ph, (W - 300, y))
        except Exception:
            pass

    draw.text((30, y), profile.get("full_name") or "Unnamed Profile", font=font(34, True), fill=indigo)
    y += 55

    for section_title, rows in biodata_lines(profile):
        draw.text((30, y), section_title, font=font(20, True), fill=gold)
        y += 28
        draw.line([(30, y), (W - 320 if y < 420 else W - 30, y)], fill=gold, width=2)
        y += 12
        for label, val in rows:
            text = f"{label}: {val}"
            draw.text((40, y), text[:110], font=font(16), fill=ink)
            y += 26
        y += 14

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    fname = f"biodata_{secure_filename(profile.get('full_name') or str(profile_id))}.png"
    return send_file(buf, mimetype="image/png", as_attachment=True, download_name=fname)


# ---------------------------------------------------------------------------
# Stats (for admin dashboard)
# ---------------------------------------------------------------------------
@app.route("/api/stats")
@login_required
def api_stats():
    db = get_db()
    total = db.execute("SELECT COUNT(*) FROM profiles").fetchone()[0]
    brides = db.execute("SELECT COUNT(*) FROM profiles WHERE profile_for = 'Bride'").fetchone()[0]
    grooms = db.execute("SELECT COUNT(*) FROM profiles WHERE profile_for = 'Groom'").fetchone()[0]
    active = db.execute("SELECT COUNT(*) FROM profiles WHERE status = 'Active'").fetchone()[0]
    pending = db.execute("SELECT COUNT(*) FROM profiles WHERE status = 'Pending'").fetchone()[0]
    return jsonify({"success": True, "total": total, "brides": brides, "grooms": grooms,
                     "active": active, "pending": pending})


@app.route("/api/reports")
@login_required
def api_reports():
    db = get_db()

    status_rows = db.execute("SELECT status, COUNT(*) AS n FROM profiles GROUP BY status").fetchall()
    by_status = {r["status"] or "Unknown": r["n"] for r in status_rows}

    gender_rows = db.execute("SELECT gender, COUNT(*) AS n FROM profiles GROUP BY gender").fetchall()
    by_gender = {r["gender"] or "Unknown": r["n"] for r in gender_rows}

    city_rows = db.execute(
        """SELECT current_city, COUNT(*) AS n FROM profiles
           WHERE current_city IS NOT NULL AND current_city != ''
           GROUP BY current_city ORDER BY n DESC LIMIT 8"""
    ).fetchall()
    top_cities = [{"label": r["current_city"], "count": r["n"]} for r in city_rows]

    caste_rows = db.execute(
        """SELECT sindhi_caste, COUNT(*) AS n FROM profiles
           WHERE sindhi_caste IS NOT NULL AND sindhi_caste != ''
           GROUP BY sindhi_caste ORDER BY n DESC LIMIT 8"""
    ).fetchall()
    top_castes = [{"label": r["sindhi_caste"], "count": r["n"]} for r in caste_rows]

    age_buckets = [("18-25", 18, 25), ("26-30", 26, 30), ("31-35", 31, 35),
                   ("36-40", 36, 40), ("41+", 41, 200)]
    age_distribution = []
    for label, lo, hi in age_buckets:
        n = db.execute("SELECT COUNT(*) FROM profiles WHERE age >= ? AND age <= ?", (lo, hi)).fetchone()[0]
        age_distribution.append({"label": label, "count": n})

    month_rows = db.execute(
        """SELECT strftime('%Y-%m', created_at) AS ym, COUNT(*) AS n FROM profiles
           GROUP BY ym ORDER BY ym DESC LIMIT 12"""
    ).fetchall()
    registrations_by_month = [{"label": r["ym"], "count": r["n"]} for r in month_rows][::-1]

    marital_rows = db.execute(
        """SELECT marital_status, COUNT(*) AS n FROM profiles
           WHERE marital_status IS NOT NULL AND marital_status != ''
           GROUP BY marital_status ORDER BY n DESC"""
    ).fetchall()
    by_marital_status = [{"label": r["marital_status"], "count": r["n"]} for r in marital_rows]

    total_matches = db.execute("SELECT COUNT(*) FROM matches").fetchone()[0]
    accepted_matches = db.execute("SELECT COUNT(*) FROM matches WHERE status = 'Accepted'").fetchone()[0]

    return jsonify({
        "success": True,
        "by_status": by_status,
        "by_gender": by_gender,
        "top_cities": top_cities,
        "top_castes": top_castes,
        "age_distribution": age_distribution,
        "registrations_by_month": registrations_by_month,
        "by_marital_status": by_marital_status,
        "total_matches": total_matches,
        "accepted_matches": accepted_matches,
    })


init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
